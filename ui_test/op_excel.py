from openpyxl import Workbook
from openpyxl import load_workbook

def newexcel():
    w=Workbook()
    ws=w.create_sheet('aaaa')
    w.save('./ee.xlsx')

def loadexcel():
    w=load_workbook('./ee.xlsx')
    ws=w.active
    print(ws['A'][0].value)

if __name__ == '__main__':
    loadexcel()