import pymysql
import logging
import os
import time
import yagmail
from openpyxl import load_workbook,Workbook
from conf.config import (email_user,
                         email_host,
                         email_password,
                         email_contents,
                         email_subject,
                         email_to,
                         sql_host,
                         sql_password,
                         sql_user,
                         sql_db,
                         sql_port,
                         excel_caseData,
                         excel_caseExpected,
                         excel_caseIsRun,
                         excel_caseNo,)

# -------------------------------------------------------------------------------
# -------------------------------------------------------------------------------
# 类名称：InsertLog
# 类的目的：写日志
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：
# 创建时间：
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
# 当前文件路径
cur_path = os.path.dirname(os.path.realpath(__file__))
# log_path是存放日志的路径
log_path = os.path.join(os.path.dirname(cur_path), 'log')
# 如果不存在这个logs文件夹，就自动创建一个
if not os.path.exists(log_path):
    os.mkdir(log_path)

class InsertLog():
    def __init__(self):
        # 文件的命名
        self.logname = os.path.join(log_path, '%s.log' % time.strftime('%Y_%m_%d'))
        self.logger = logging.getLogger()
        self.logger.setLevel(logging.DEBUG)
        # 日志输出格式
        self.formatter = logging.Formatter(
            '[%(asctime)s - %(funcName)s line: %(lineno)3d] - %(levelname)s: %(message)s')

    def __console(self, level, message):
        # 创建一个FileHandler，用于写到本地
        # fh = logging.FileHandler(self.logname, 'a')  # 追加模式  这个是python2的
        fh = logging.FileHandler(self.logname, 'a', encoding='utf-8')  # 这个是python3的
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(self.formatter)
        self.logger.addHandler(fh)

        # 创建一个StreamHandler,用于输出到控制台
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)
        ch.setFormatter(self.formatter)
        self.logger.addHandler(ch)

        if level == 'info':
            self.logger.info(message)
        elif level == 'debug':
            self.logger.debug(message)
        elif level == 'warning':
            self.logger.warning(message)
        elif level == 'error':
            self.logger.error(message)
        # 这两行代码是为了避免日志输出重复问题
        self.logger.removeHandler(ch)
        self.logger.removeHandler(fh)
        # 关闭打开的文件
        fh.close()

    def debug(self, message):
        self.__console('debug', message)

    def info(self, message):
        self.__console('info', message)

    def warning(self, message):
        self.__console('warning', message)

    def error(self, message):
        self.__console('error', message)

# -------------------------------------------------------------------------------
# 类名称：connectSql
# 类的目的：连接数据库,处理sql
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：
# 创建时间：
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------

def connectSql(sql='',
               user=sql_user,
               password=sql_password,
               host=sql_host,
               port=sql_port,
               db=sql_db):
    try:

        # with pymysql.connect(user=user,
        #                      password=password,
        #                      host=host,
        #                      port=port,
        #                      db=db) as my:
        #     my.execute(sql)
        #     result=my.fetchone()
        # -------第二种表达式-------
        my = pymysql.connect(user=user, password=password, host=host, port=port, db=db)
        mm=my.cursor()
        mm.execute(sql)
        result=mm.fetchone()
        my.close()
        mm.close()

        return result

    except BaseException as e:
        log=InsertLog()
        log.error('sql执行失败······')
        log.error(e)

# -------------------------------------------------------------------------------
# 类名称：sendEmail
# 类的目的：发送邮件
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：
# 创建时间：
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
def sendEmail(attachments='',
              user=email_user,
              password=email_password,
              host=email_host,
              to=email_to,
              subject=email_subject,
              contents=email_contents,
              cc=email_user
              ):
    mail=yagmail.SMTP(user=user,password=password,host=host)
    mail.send(to=to,subject=subject,contents=contents,attachments=attachments,cc=cc)

# -------------------------------------------------------------------------------
# 类名称：
# 类的目的：操作Excel文档
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：
# 创建时间：
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------

class DoExcel():
    def __init__(self,sheetname='Login_cases',filepath='D:/xmart/cases/cases.xlsx'):
        self.wb =load_workbook(filepath)
        self.ws = self.wb[sheetname]

    def get_sheets_name(self):
        # 获取所有表单名称
        return self.wb.get_sheet_names()

    def get_index_by_data(self,col,value):
        #  根据内容获取下标
        data=self.ws[col]
        index=[]
        num=0
        for i in data:
            if value in i.value:
                index.append(num)
            num += 1
        return index

    def get_data_by_index(self,col,indexs):
        #  根据下标获取内容
        data = self.ws[col]
        is_run=[]
        for j in indexs:
            is_run.append(data[j].value)
        return is_run

    def get_is_run_name(self):
        # 获取需要执行的用例编号
        index=self.get_index_by_data(excel_caseIsRun,'Y')
        name=self.get_data_by_index(excel_caseNo,index)
        return name

    def get_test_data(self,casename):
        # 根据用例编号获取测试数据,根据key值返回具体数据
        index=self.get_index_by_data(excel_caseNo,casename)
        name=self.get_data_by_index(excel_caseData,index)
        data_dict=eval(name[0])
        return data_dict

if __name__ == '__main__':
    # username = '16712341234'
    # sql = "delete a from xsmart_users a , (select max(id) as m_id from xsmart_users) b where a.id=b.m_id"
    # # sql='select max(id) from xsmart_users'
    # a=connectSql(sql)
    # print(a)
    # sendEmail('./report/report190527142535.html')
    data=DoExcel('Vip_cases')
    aa=data.get_test_data('test_login_001')
    # aa=data.get_sheets_name()
    print(aa)